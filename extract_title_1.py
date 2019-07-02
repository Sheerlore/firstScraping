import sys
import requests
from bs4 import BeautifulSoup
import datetime
import openpyxl as px
import random

today = datetime.date.today()
time = datetime.datetime.now().strftime("%H:%M:%S")
date = today.isoformat()
numran = str(random.randint(10000, 199999))

list_keywd = ['python']
#検索数
search_num = 10
#GoogleのURL
search_URL = 'https://www.google.co.jp/search?num='+ str(search_num)
#User-Agent
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36 "}

#Scraping用関数
def Google_search_scraping():

    #Excelファイル出力
    #保存用Workbook作成
    workbook = px.Workbook()
    sheet = workbook.active
    sheet.title = date

    #Googleに接続
    request_Google = requests.get(search_URL, params={'q':' '.join(list_keywd) },headers = headers)
    request_Google.raise_for_status()

    #BeautifulSoupで掲載サイトのURLを取得
    bs4_Google = BeautifulSoup(request_Google.text,"html.parser")


    #scriptタグを削除
    for script in bs4_Google.find_all('script',src = False):
        script.decompose()
    #styleタグを削除
    for style in bs4_Google.find_all('style',src = False):
        style.decompose()

    #print(bs4_Google.prettify())

    #検索結果のタイトルを取得
    title = bs4_Google.find_all('h3',{'class':'LC20lb'})
    url = bs4_Google.find_all('cite',{'class':'iUh30'})

    #Excelの設定
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 50
    sheet['A1'].value = '順位'
    sheet['B1'].value = 'タイトル'
    sheet['C1'].value = 'URL'

    #出力
    for (tl,ul) in zip(title, url) :
        print(" ・ " + tl.text)
        print("     <"+ ul.text+">")
        print("\n")

    for i, (tl,ul) in enumerate(zip(title,url)):
        sheet.cell(row=i+2,column=1,value=i+1)
        sheet.cell(row=i+2,column=2,value=tl.text)
        sheet.cell(row=i+2,column=3).hyperlink = ul.text



    #Excel保存
    workbook.save(f'{date}--{numran}.xlsx')


if __name__ == '__main__':
    Google_search_scraping()
    print(date)
    print(time)
